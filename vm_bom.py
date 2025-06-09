#!/usr/bin/env python3
"""
Simplified VM BOM Generator with Oracle Cloud Pricing
Reads VM inventory CSV and generates cost report
Fixed to work with actual CSV column format
FIXED: Added proper rounding for all values
"""

import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import List
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

@dataclass
class VMSpec:
    """VM specification from CSV"""
    vm_name: str
    os_config: str
    cpu_cpus: int
    mem_size_gb: float
    disk_total_capacity_gb: float
    annotation: str
    powerstate: str = ""

@dataclass
class BOMLine:
    """Bill of Materials line item"""
    vm_name: str
    component_type: str
    description: str
    quantity: float
    unit: str
    unit_price: float
    total_cost: float

class VMBOMGenerator:
    def __init__(self, debug=False):
        # Oracle Cloud pricing data (EUR)
        self.pricing_data = {
            "compute": {"description": "OCPU per hour", "unit_price": 0.0279, "unit": "OCPU/hour"},
            "memory": {"description": "Memory GB per hour", "unit_price": 0.00186, "unit": "GB/hour"},
            "storage": {"description": "Block Volume Storage", "unit_price": 0.023715, "unit": "GB/month"},
            "storage_vpu": {"description": "Block Volume VPUs", "unit_price": 0.001581, "unit": "VPU/month"},
            "windows": {"description": "Windows Server License", "unit_price": 0.08556, "unit": "OCPU/hour"}
        }
        
        self.hours_per_month = 744  # 31 days * 24 hours
        self.debug = debug
    
    def debug_print(self, message):
        """Print debug message if debug mode is enabled"""
        if self.debug:
            print(f"[DEBUG] {message}")
    
    def detect_os_type(self, os_config: str) -> str:
        """Detect OS type from os_config string"""
        os_lower = os_config.lower()
        if 'windows' in os_lower or 'microsoft' in os_lower:
            return 'windows'
        elif any(linux in os_lower for linux in ['ubuntu', 'centos', 'oracle linux', 'debian', 'suse', 'linux']):
            return 'linux'
        return 'other'
    
    def calculate_ocpu_count(self, cpu_cpus: int) -> int:
        """Calculate OCPU count (1 OCPU = 2 vCPUs, minimum 1 OCPU)"""
        if cpu_cpus <= 0:
            return 0
        if cpu_cpus == 1:
            ocpu = 1
        else:
            ocpu = (cpu_cpus + 1) // 2  # Round up
        
        self.debug_print(f"vCPUs: {cpu_cpus} -> OCPUs: {ocpu}")
        return ocpu
    
    def calculate_vm_pricing(self, vm_spec: VMSpec) -> List[BOMLine]:
        """Calculate pricing for a single VM"""
        bom_lines = []
        os_type = self.detect_os_type(vm_spec.os_config)
        ocpu_count = self.calculate_ocpu_count(vm_spec.cpu_cpus)
        
        self.debug_print(f"\nCalculating pricing for VM: {vm_spec.vm_name}")
        self.debug_print(f"OS: {vm_spec.os_config} (detected as: {os_type})")
        self.debug_print(f"vCPUs: {vm_spec.cpu_cpus}, OCPUs: {ocpu_count}")
        self.debug_print(f"Memory: {vm_spec.mem_size_gb} GB")
        self.debug_print(f"Storage: {vm_spec.disk_total_capacity_gb} GB")
        
        # Skip VMs with no resources or powered off
        if (ocpu_count == 0 and vm_spec.mem_size_gb == 0 and vm_spec.disk_total_capacity_gb == 0) or vm_spec.powerstate.lower() != 'poweredon':
            if vm_spec.powerstate.lower() != 'poweredon':
                self.debug_print(f"Skipping {vm_spec.vm_name} - not powered on ({vm_spec.powerstate})")
            return bom_lines
        
        # 1. Compute - OCPU
        if ocpu_count > 0:
            ocpu_hourly = self.pricing_data["compute"]["unit_price"]
            ocpu_monthly = ocpu_hourly * self.hours_per_month
            ocpu_cost = ocpu_count * ocpu_monthly
            
            bom_lines.append(BOMLine(
                vm_name=vm_spec.vm_name,
                component_type="Compute",
                description=f"OCPU ({ocpu_count} OCPU for {vm_spec.cpu_cpus} vCPU)",
                quantity=float(ocpu_count),  # Ensure float for consistency
                unit="OCPU",
                unit_price=round(ocpu_monthly, 4),
                total_cost=round(ocpu_cost, 2)
            ))
        
        # 2. Memory
        if vm_spec.mem_size_gb > 0:
            mem_hourly = self.pricing_data["memory"]["unit_price"]
            mem_monthly = mem_hourly * self.hours_per_month
            memory_cost = vm_spec.mem_size_gb * mem_monthly
            
            bom_lines.append(BOMLine(
                vm_name=vm_spec.vm_name,
                component_type="Memory",
                description=f"Memory ({vm_spec.mem_size_gb:.1f} GB)",
                quantity=round(vm_spec.mem_size_gb, 1),
                unit="GB",
                unit_price=round(mem_monthly, 4),
                total_cost=round(memory_cost, 2)
            ))
        
        # 3. Block Storage
        if vm_spec.disk_total_capacity_gb > 0:
            storage_monthly = self.pricing_data["storage"]["unit_price"]
            storage_cost = vm_spec.disk_total_capacity_gb * storage_monthly
            
            bom_lines.append(BOMLine(
                vm_name=vm_spec.vm_name,
                component_type="Storage",
                description=f"Block Volume Storage ({vm_spec.disk_total_capacity_gb:.1f} GB)",
                quantity=round(vm_spec.disk_total_capacity_gb, 1),
                unit="GB",
                unit_price=round(storage_monthly, 4),
                total_cost=round(storage_cost, 2)
            ))
            
            # 4. Storage VPUs (10 VPUs per GB of storage) - FIX: Proper rounding
            vpu_count = vm_spec.disk_total_capacity_gb * 10
            vpu_monthly = self.pricing_data["storage_vpu"]["unit_price"]
            vpu_cost = vpu_count * vpu_monthly
            
            bom_lines.append(BOMLine(
                vm_name=vm_spec.vm_name,
                component_type="Storage Performance",
                description=f"Block Volume VPUs ({round(vpu_count, 1)} VPUs)",
                quantity=round(vpu_count, 1),  # FIX: Round to 1 decimal place
                unit="VPU",
                unit_price=round(vpu_monthly, 4),
                total_cost=round(vpu_cost, 2)
            ))
        
        # 5. Windows Licensing
        if os_type == 'windows' and ocpu_count > 0:
            windows_hourly = self.pricing_data["windows"]["unit_price"]
            windows_monthly = windows_hourly * self.hours_per_month
            windows_cost = ocpu_count * windows_monthly
            
            bom_lines.append(BOMLine(
                vm_name=vm_spec.vm_name,
                component_type="OS License",
                description=f"Windows Server License ({ocpu_count} OCPU)",
                quantity=float(ocpu_count),  # Ensure float for consistency
                unit="OCPU",
                unit_price=round(windows_monthly, 4),
                total_cost=round(windows_cost, 2)
            ))
        
        return bom_lines
    
    def read_vm_csv(self, csv_file_path: str) -> List[VMSpec]:
        """Read VM specifications from CSV file with flexible column mapping"""
        vm_specs = []
        
        try:
            with open(csv_file_path, 'r', encoding='utf-8-sig') as file:
                reader = csv.DictReader(file)
                
                # Clean column names
                if reader.fieldnames:
                    reader.fieldnames = [field.strip() for field in reader.fieldnames]
                
                print(f"CSV columns found: {reader.fieldnames}")
                
                # Map actual column names to expected names
                column_mapping = {}
                for field in reader.fieldnames:
                    field_lower = field.lower()
                    if 'cpu_vm' in field_lower or field_lower == 'vm name' or field_lower == 'vm_name':
                        column_mapping['vm_name'] = field
                    elif 'cpu_os' in field_lower or 'os according' in field_lower:
                        column_mapping['os_config'] = field
                    elif 'cpu_cpus' in field_lower or field_lower == 'vcpu' or field_lower == 'cpus':
                        column_mapping['cpu_cpus'] = field
                    elif 'memory_size' in field_lower or 'mem_size' in field_lower or field_lower == 'memory gb':
                        column_mapping['mem_size_gb'] = field
                    elif 'disk_capacity' in field_lower or 'disk_total' in field_lower or field_lower == 'disk gb':
                        column_mapping['disk_total_capacity_gb'] = field
                    elif 'annotation' in field_lower or 'notes' in field_lower:
                        column_mapping['annotation'] = field
                    elif 'powerstate' in field_lower or 'power_state' in field_lower:
                        column_mapping['powerstate'] = field
                
                print(f"Column mapping: {column_mapping}")
                
                # Check required columns
                required = ['vm_name', 'os_config', 'cpu_cpus', 'mem_size_gb', 'disk_total_capacity_gb']
                missing = [col for col in required if col not in column_mapping]
                
                if missing:
                    print(f"ERROR: Cannot find columns for: {missing}")
                    print("Available columns:", reader.fieldnames)
                    return []
                
                # Process rows
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Skip empty rows
                        vm_name_col = column_mapping.get('vm_name')
                        if not vm_name_col or not row.get(vm_name_col, '').strip():
                            continue
                        
                        vm_name = row[vm_name_col].strip()
                        os_config = row.get(column_mapping.get('os_config', ''), '').strip()
                        annotation = row.get(column_mapping.get('annotation', ''), '').strip()
                        powerstate = row.get(column_mapping.get('powerstate', ''), 'poweredOn').strip()
                        
                        try:
                            cpu_cpus = int(float(row.get(column_mapping.get('cpu_cpus', ''), 0)))
                        except (ValueError, TypeError):
                            cpu_cpus = 0
                        
                        try:
                            mem_size_gb = float(row.get(column_mapping.get('mem_size_gb', ''), 0))
                        except (ValueError, TypeError):
                            mem_size_gb = 0.0
                        
                        try:
                            disk_total_capacity_gb = float(row.get(column_mapping.get('disk_total_capacity_gb', ''), 0))
                        except (ValueError, TypeError):
                            disk_total_capacity_gb = 0.0
                        
                        # Skip VMs with zero resources
                        if cpu_cpus == 0 and mem_size_gb == 0 and disk_total_capacity_gb == 0:
                            continue
                        
                        vm_specs.append(VMSpec(
                            vm_name=vm_name,
                            os_config=os_config,
                            cpu_cpus=cpu_cpus,
                            mem_size_gb=mem_size_gb,
                            disk_total_capacity_gb=disk_total_capacity_gb,
                            annotation=annotation,
                            powerstate=powerstate
                        ))
                        
                    except Exception as e:
                        print(f"Warning: Error processing row {row_num}: {e}")
                        continue
                
                print(f"Successfully loaded {len(vm_specs)} VMs")
                    
        except FileNotFoundError:
            print(f"Error: File '{csv_file_path}' not found")
        except Exception as e:
            print(f"Error reading CSV: {e}")
        
        return vm_specs
    
    def generate_cost_report(self, vm_specs: List[VMSpec]) -> str:
        """Generate complete cost report"""
        if not vm_specs:
            return "No VMs to process"
        
        all_bom_lines = []
        vm_summaries = {}
        powered_off_vms = []
        
        # Calculate pricing for each VM
        for vm_spec in vm_specs:
            if vm_spec.powerstate.lower() != 'poweredon':
                powered_off_vms.append(vm_spec)
                continue
                
            bom_lines = self.calculate_vm_pricing(vm_spec)
            if bom_lines:  # Only include VMs with costs
                all_bom_lines.extend(bom_lines)
                vm_total = sum(line.total_cost for line in bom_lines)
                vm_summaries[vm_spec.vm_name] = {
                    'spec': vm_spec,
                    'monthly_cost': round(vm_total, 2),  # FIX: Round monthly cost
                    'annual_cost': round(vm_total * 12, 2),  # FIX: Round annual cost
                    'bom_lines': bom_lines
                }
        
        if not vm_summaries:
            return "No valid powered-on VMs with pricing found"
        
        # Generate report
        report = []
        report.append("=" * 120)
        report.append("VIRTUAL MACHINE COST ANALYSIS REPORT")
        report.append("Oracle Cloud Infrastructure Pricing")
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("Pricing in EUR - 24/7 operation (744 hours/month)")
        report.append("=" * 120)
        report.append("")
        
        # Executive Summary
        total_monthly = round(sum(s['monthly_cost'] for s in vm_summaries.values()), 2)  # FIX: Round total
        total_annual = round(total_monthly * 12, 2)  # FIX: Round total
        
        report.append("EXECUTIVE SUMMARY")
        report.append("-" * 60)
        report.append(f"Total VMs Analyzed: {len(vm_summaries)} (powered on)")
        if powered_off_vms:
            report.append(f"VMs Excluded (powered off): {len(powered_off_vms)}")
        report.append(f"Monthly Total Cost: €{total_monthly:,.2f}")
        report.append(f"Annual Total Cost: €{total_annual:,.2f}")
        report.append(f"Average Cost per VM: €{total_monthly/len(vm_summaries):,.2f}/month")
        report.append("")
        
        # VM Details Table
        report.append("DETAILED VM COST BREAKDOWN")
        report.append("-" * 120)
        report.append(f"{'VM Name':<25} {'OS Type':<12} {'vCPU':<5} {'RAM':<8} {'Disk':<8} {'Monthly':<10} {'Annual':<12} {'Notes':<30}")
        report.append("-" * 120)
        
        sorted_vms = sorted(vm_summaries.items(), key=lambda x: x[1]['monthly_cost'], reverse=True)
        for vm_name, summary in sorted_vms:
            spec = summary['spec']
            os_type = self.detect_os_type(spec.os_config).title()
            notes = spec.annotation[:28] + ".." if len(spec.annotation) > 30 else spec.annotation
            report.append(f"{vm_name:<25} {os_type:<12} {spec.cpu_cpus:<5} {spec.mem_size_gb:<8.1f} {spec.disk_total_capacity_gb:<8.1f} €{summary['monthly_cost']:<9.2f} €{summary['annual_cost']:<11,.2f} {notes:<30}")
        
        report.append("-" * 120)
        report.append(f"{'TOTAL':<25} {'':<12} {'':<5} {'':<8} {'':<8} €{total_monthly:<9.2f} €{total_annual:<11,.2f}")
        report.append("")
        
        # Powered off VMs summary
        if powered_off_vms:
            report.append("POWERED OFF VMs (Not included in cost calculation)")
            report.append("-" * 80)
            for vm in powered_off_vms:
                os_type = self.detect_os_type(vm.os_config).title()
                notes = vm.annotation[:40] + ".." if len(vm.annotation) > 42 else vm.annotation
                report.append(f"{vm.vm_name:<25} {os_type:<12} {vm.cpu_cpus:<5} {vm.mem_size_gb:<8.1f} {vm.disk_total_capacity_gb:<8.1f} {notes:<42}")
            report.append("")
        
        # Component Cost Breakdown
        report.append("COST BREAKDOWN BY COMPONENT TYPE")
        report.append("-" * 70)
        
        component_totals = {}
        for line in all_bom_lines:
            if line.component_type not in component_totals:
                component_totals[line.component_type] = 0
            component_totals[line.component_type] += line.total_cost
        
        for component, cost in sorted(component_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / total_monthly) * 100
            report.append(f"{component:<30} €{cost:>10.2f}/month ({percentage:>5.1f}%)")
        
        report.append("")
        
        # Detailed Component Analysis
        report.append("DETAILED COMPONENT ANALYSIS")
        report.append("-" * 120)
        report.append(f"{'VM Name':<25} {'Component':<20} {'Description':<35} {'Qty':<8} {'Unit Price':<12} {'Total':<10}")
        report.append("-" * 120)
        
        for vm_name, summary in sorted_vms:
            for i, line in enumerate(summary['bom_lines']):
                vm_display = vm_name if i == 0 else ""
                report.append(f"{vm_display:<25} {line.component_type:<20} {line.description:<35} {line.quantity:<8.1f} €{line.unit_price:<11.4f} €{line.total_cost:<9.2f}")
            if summary['bom_lines']:
                report.append("-" * 120)
        
        return "\n".join(report)
    
    def export_detailed_analysis_to_excel(self, vm_specs: List[VMSpec], output_file: str):
        """Export detailed component analysis to Excel file"""
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library not installed. Install with: pip install openpyxl")
            return
        
        if not vm_specs:
            print("No VMs to export")
            return
        
        # Calculate all BOM lines
        all_bom_lines = []
        vm_summaries = {}
        powered_off_vms = []
        
        for vm_spec in vm_specs:
            if vm_spec.powerstate.lower() != 'poweredon':
                powered_off_vms.append(vm_spec)
                continue
                
            bom_lines = self.calculate_vm_pricing(vm_spec)
            if bom_lines:
                all_bom_lines.extend(bom_lines)
                vm_total = sum(line.total_cost for line in bom_lines)
                vm_summaries[vm_spec.vm_name] = {
                    'spec': vm_spec,
                    'monthly_cost': round(vm_total, 2),  # FIX: Round for Excel too
                    'annual_cost': round(vm_total * 12, 2),  # FIX: Round for Excel too
                    'bom_lines': bom_lines
                }
        
        if not vm_summaries:
            print("No valid VMs with pricing found")
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Cost Summary"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Data styling
        currency_format = '€#,##0.00'
        number_format = '#,##0.0'
        
        # Summary sheet headers
        summary_headers = ["VM Name", "OS Type", "vCPU", "RAM (GB)", "Disk (GB)", "Monthly Cost", "Annual Cost", "Notes"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Summary data
        total_monthly = round(sum(s['monthly_cost'] for s in vm_summaries.values()), 2)  # FIX: Round total
        total_annual = round(total_monthly * 12, 2)  # FIX: Round total
        
        sorted_vms = sorted(vm_summaries.items(), key=lambda x: x[1]['monthly_cost'], reverse=True)
        row = 2
        for vm_name, summary in sorted_vms:
            spec = summary['spec']
            os_type = self.detect_os_type(spec.os_config).title()
            
            ws_summary.cell(row=row, column=1, value=vm_name)
            ws_summary.cell(row=row, column=2, value=os_type)
            ws_summary.cell(row=row, column=3, value=spec.cpu_cpus)
            ws_summary.cell(row=row, column=4, value=spec.mem_size_gb).number_format = number_format
            ws_summary.cell(row=row, column=5, value=spec.disk_total_capacity_gb).number_format = number_format
            ws_summary.cell(row=row, column=6, value=summary['monthly_cost']).number_format = currency_format
            ws_summary.cell(row=row, column=7, value=summary['annual_cost']).number_format = currency_format
            ws_summary.cell(row=row, column=8, value=spec.annotation)
            row += 1
        
        # Total row
        ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws_summary.cell(row=row, column=6, value=total_monthly).number_format = currency_format
        ws_summary.cell(row=row, column=6).font = Font(bold=True)
        ws_summary.cell(row=row, column=7, value=total_annual).number_format = currency_format
        ws_summary.cell(row=row, column=7).font = Font(bold=True)
        
        # Powered off VMs sheet
        if powered_off_vms:
            ws_powered_off = wb.create_sheet("Powered Off VMs")
            
            headers = ["VM Name", "OS Type", "vCPU", "RAM (GB)", "Disk (GB)", "Power State", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = ws_powered_off.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row = 2
            for vm in powered_off_vms:
                os_type = self.detect_os_type(vm.os_config).title()
                ws_powered_off.cell(row=row, column=1, value=vm.vm_name)
                ws_powered_off.cell(row=row, column=2, value=os_type)
                ws_powered_off.cell(row=row, column=3, value=vm.cpu_cpus)
                ws_powered_off.cell(row=row, column=4, value=vm.mem_size_gb).number_format = number_format
                ws_powered_off.cell(row=row, column=5, value=vm.disk_total_capacity_gb).number_format = number_format
                ws_powered_off.cell(row=row, column=6, value=vm.powerstate)
                ws_powered_off.cell(row=row, column=7, value=vm.annotation)
                row += 1
        
        # Detailed analysis sheet
        ws_detail = wb.create_sheet("Detailed Analysis")
        
        # Detailed headers
        detail_headers = ["VM Name", "Component Type", "Description", "Quantity", "Unit", "Unit Price", "Monthly Cost"]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Detailed data
        row = 2
        for vm_name, summary in sorted_vms:
            vm_start_row = row
            for line in summary['bom_lines']:
                ws_detail.cell(row=row, column=1, value=vm_name)
                ws_detail.cell(row=row, column=2, value=line.component_type)
                ws_detail.cell(row=row, column=3, value=line.description)
                ws_detail.cell(row=row, column=4, value=line.quantity).number_format = number_format
                ws_detail.cell(row=row, column=5, value=line.unit)
                ws_detail.cell(row=row, column=6, value=line.unit_price).number_format = currency_format
                ws_detail.cell(row=row, column=7, value=line.total_cost).number_format = currency_format
                row += 1
            
            # VM subtotal
            if summary['bom_lines']:
                ws_detail.cell(row=row, column=1, value=f"{vm_name} Subtotal").font = Font(italic=True)
                ws_detail.cell(row=row, column=7, value=summary['monthly_cost']).number_format = currency_format
                ws_detail.cell(row=row, column=7).font = Font(italic=True)
                row += 1
                row += 1  # Empty row for separation
        
        # Component breakdown sheet
        ws_components = wb.create_sheet("Component Breakdown")
        
        # Component breakdown headers
        comp_headers = ["Component Type", "Monthly Cost", "Percentage"]
        for col, header in enumerate(comp_headers, 1):
            cell = ws_components.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Component data
        component_totals = {}
        for line in all_bom_lines:
            if line.component_type not in component_totals:
                component_totals[line.component_type] = 0
            component_totals[line.component_type] += line.total_cost
        
        row = 2
        for component, cost in sorted(component_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (cost / total_monthly) * 100
            ws_components.cell(row=row, column=1, value=component)
            ws_components.cell(row=row, column=2, value=cost).number_format = currency_format
            ws_components.cell(row=row, column=3, value=percentage/100).number_format = '0.0%'
            row += 1
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            wb.save(output_file)
            print(f"Detailed analysis exported to Excel: {output_file}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage: python vm_bom.py <input_csv_file> [--debug] [--excel]")
        print("Example: python vm_bom.py vm_inventory.csv")
        print("         python vm_bom.py vm_inventory.csv --debug")
        print("         python vm_bom.py vm_inventory.csv --excel")
        print("         python vm_bom.py vm_inventory.csv --debug --excel")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    debug_mode = "--debug" in sys.argv
    export_excel = "--excel" in sys.argv
    
    if export_excel and not EXCEL_AVAILABLE:
        print("Warning: Excel export requested but openpyxl not installed.")
        print("Install with: pip install openpyxl")
        print("Continuing with console report only...")
        export_excel = False
    
    generator = VMBOMGenerator(debug=debug_mode)
    
    print(f"Reading VM specifications from: {csv_file}")
    vm_specs = generator.read_vm_csv(csv_file)
    
    if not vm_specs:
        print("No valid VM specifications found.")
        sys.exit(1)
    
    print(f"\nGenerating cost analysis for {len(vm_specs)} VMs...")
    
    # Generate and display complete report
    report = generator.generate_cost_report(vm_specs)
    print("\n" + report)
    
    # Export to Excel if requested
    if export_excel:
        output_excel = csv_file.replace('.csv', '_detailed_analysis.xlsx')
        generator.export_detailed_analysis_to_excel(vm_specs, output_excel)

if __name__ == "__main__":
    main()